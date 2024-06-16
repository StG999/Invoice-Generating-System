import openpyxl
from num2words import num2words
from win32com import client 
import pythoncom
import os

def getGrandTotalString(grandTotal):
    grandTotal = float(grandTotal)
    s = ''
    try:
        s = num2words(grandTotal, to='currency', lang='en_IN')
    except NotImplementedError:
        s = num2words(grandTotal, lang='en')
    
    s = s.title()
    s = s.replace('And', 'and')
    if grandTotal < 2.0:
        s = s.replace('Euro', 'Rupee')
    else: 
        s = s.replace('Euro', 'Rupees')
    s = s.replace('Cents', 'Paise')
    return s
    
def setItems(ws, items):
    for i, item in enumerate(items):
        ws[f'H{i+11}'] = item['name']
        ws[f'K{i+11}'] = item['packaging']
        ws[f'L{i+11}'] = item['quantity']
        ws[f'M{i+11}'] = item['rate']

def createNewBill(invoiceNumber, date, customerName, customerAddress, items, grandTotal):
    wb = openpyxl.load_workbook('template.xlsx')
    ws = wb['Sheet1']

    ws['N7'] = date
    ws['J7'] = f": {invoiceNumber}"
    ws['J8'] = f": {customerName}"
    ws['J9'] = f": {customerAddress}"
    
    setItems(ws, items)
    stringGrandTotal = getGrandTotalString(grandTotal)
    ws['I29'] = stringGrandTotal
    wb.save('newBill.xlsx')

def convertToPdf():
    # Open Microsoft Excel 
    pythoncom.CoInitialize()
    path = os.getcwd()
    excel = client.Dispatch("\Excel.Application") 
    
    # Read Excel File 
    sheets = excel.Workbooks.Open(path + '/newBill.xlsx') 
    work_sheets = sheets.Worksheets[0] 
    
    # Convert into PDF File 
    work_sheets.ExportAsFixedFormat(0, path + '/newBill.pdf') 
    sheets.Close()